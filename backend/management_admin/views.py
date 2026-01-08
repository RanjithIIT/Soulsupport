"""
Views for management_admin app - API layer for App 2
"""
import random
import string
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
from io import BytesIO
from django.http import HttpResponse
from django.core.files.base import ContentFile
from .models import File, Department, Teacher, Student, DashboardStats, NewAdmission, Examination_management, Fee, PaymentHistory, Bus, BusStop, BusStopStudent, Event, Award, AwardCertificate, CampusFeature, Activity, Gallery, GalleryImage
from super_admin.models import School
from .serializers import (
    FileSerializer,
    DepartmentSerializer,
    TeacherSerializer,
    StudentSerializer,
    DashboardStatsSerializer,
    NewAdmissionSerializer,
    ExaminationManagementSerializer,
    FeeSerializer,
    BusSerializer,
    BusStopSerializer,
    BusStopStudentSerializer,
    EventSerializer,
    AwardSerializer,
    CampusFeatureSerializer,
    ActivitySerializer,
    GallerySerializer,
    GalleryImageSerializer
)


from main_login.permissions import IsManagementAdmin
from main_login.mixins import SchoolFilterMixin
from main_login.utils import get_user_school_id
from django.conf import settings



class FileViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for File uploads (profile photos)"""
    queryset = File.objects.all()
    serializer_class = FileSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['file_type', 'school_id']
    search_fields = ['file_name']
    ordering_fields = ['created_at']
    ordering = ['-created_at']
    
    def perform_create(self, serializer):
        """Set uploaded_by and school_id when creating file"""
        school_id = get_user_school_id(self.request.user)
        serializer.save(
            uploaded_by=self.request.user,
            school_id=school_id
        )



# Teacher ViewSet


class TeacherViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Teacher management"""
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['department', 'is_active']
    search_fields = ['user__first_name', 'user__last_name', 'employee_no', 'email', 'first_name', 'last_name']
    ordering_fields = ['joining_date', 'created_at']
    ordering = ['-created_at']

    def get_permissions(self):
        """Require authentication for list/retrieve to ensure school filtering works"""
        # Changed: Require authentication for list/retrieve to enable school filtering
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]  # Changed from AllowAny() to ensure school filtering
        if self.action in ['create', 'destroy']:
            return [AllowAny()]  # Keep AllowAny for create/destroy if needed
        return [IsAuthenticated(), IsManagementAdmin()]
    
    def perform_create(self, serializer):
        """Override to ensure school_id is set correctly when creating teachers"""
        school_id = self.get_school_id()
        
        # Check if user is super admin
        if hasattr(self.request.user, 'role') and self.request.user.role:
            if self.request.user.role.name == 'super_admin':
                # Super admin can set school_id manually or leave it
                # If school_id is provided in data, use it; otherwise let it be set from department
                super().perform_create(serializer)
                # After save, ensure school_id and school_name are set from department if not already set
                teacher = serializer.instance
                if teacher and teacher.department_id:
                    from management_admin.models import Department
                    try:
                        department = Department.objects.select_related('school').get(pk=teacher.department_id)
                        if department.school:
                            department_school_id = department.school.school_id
                            department_school_name = department.school.school_name
                            needs_save = False
                            # Always update school_id if it's different
                            if not teacher.school_id or teacher.school_id != department_school_id:
                                teacher.school_id = department_school_id
                                needs_save = True
                            # Always update school_name if it's missing or different (even if school_id is already set)
                            if not teacher.school_name or teacher.school_name != department_school_name:
                                teacher.school_name = department_school_name
                                needs_save = True
                            if needs_save:
                                teacher.save(update_fields=['school_id', 'school_name'])
                    except Department.DoesNotExist:
                        pass
                return
        
        # For non-super-admin users, automatically set school_id
        serializer.save()
        
        # After save, ensure school_id and school_name are set
        teacher = serializer.instance
        if teacher:
            # First, try to get school_id and school_name from department's school
            if teacher.department:
                # Refresh department to ensure school relationship is loaded
                from management_admin.models import Department
                try:
                    department = Department.objects.select_related('school').get(pk=teacher.department_id)
                    if department.school:
                        department_school_id = department.school.school_id
                        department_school_name = department.school.school_name
                        needs_save = False
                        # Always update school_id if it's different
                        if not teacher.school_id or teacher.school_id != department_school_id:
                            teacher.school_id = department_school_id
                            needs_save = True
                        # Always update school_name if it's missing or different (even if school_id is already set)
                        if not teacher.school_name or teacher.school_name != department_school_name:
                            teacher.school_name = department_school_name
                            needs_save = True
                        if needs_save:
                            teacher.save(update_fields=['school_id', 'school_name'])
                except Department.DoesNotExist:
                    pass
            # If no department or department has no school, use school_id from user context
            elif school_id and not teacher.school_id:
                from super_admin.models import School
                try:
                    school = School.objects.get(school_id=school_id)
                    teacher.school_id = school_id
                    teacher.school_name = school.school_name
                    teacher.save(update_fields=['school_id', 'school_name'])
                except School.DoesNotExist:
                    pass
    
    def create(self, request, *args, **kwargs):
        """
        Override create to allow creation even if school_id is not found initially.
        school_id will be auto-populated from department when teacher is saved.
        This completely bypasses the SchoolFilterMixin.create() check.
        Also handles profile photo upload.
        """
        # Handle profile photo upload if present
        profile_photo_file = request.FILES.get('profile_photo')
        profile_photo_url = None
        
        if profile_photo_file:
            # Save file and get the URL/path
            from django.core.files.storage import default_storage
            from django.utils import timezone
            import os
            
            # Generate filename with timestamp to avoid conflicts
            file_ext = os.path.splitext(profile_photo_file.name)[1]
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            filename = f'profile_photos/{timestamp}_{profile_photo_file.name}'
            
            # Save file using default storage
            saved_path = default_storage.save(filename, profile_photo_file)
            # Get the URL for the saved file
            profile_photo_url = default_storage.url(saved_path)
        
        # Prepare data for serializer
        data = request.data.copy()
        if 'profile_photo' in data:
            if profile_photo_file:
                # Remove the file from data, we'll add the URL instead
                del data['profile_photo']
            elif isinstance(data['profile_photo'], str):
                # If profile_photo is already a string (URL), use it
                profile_photo_url = data['profile_photo']
        
        if profile_photo_url:
            data['profile_photo'] = profile_photo_url
        
        # Get serializer
        serializer = self.get_serializer(data=data)
        
        # Validate and log errors if validation fails
        if not serializer.is_valid():
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Teacher creation validation failed: {serializer.errors}")
            logger.error(f"Request data: {data}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Call perform_create which will handle school_id assignment
        # This allows creation even if school_id is not found (will be set from department)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    def get_queryset(self):
        """
        Override to filter by school_id using SchoolFilterMixin logic.
        Now all list/retrieve requests are authenticated, so filtering will always work.
        """
        # Call SchoolFilterMixin's get_queryset to get proper filtering
        queryset = super(SchoolFilterMixin, self).get_queryset()
        
        # All requests should be authenticated now (due to permission change above)
        # But keep the check for safety
        if not self.request.user.is_authenticated:
            return queryset.none()  # Changed: Return empty instead of all teachers
        
        # Check if user is super admin
        if hasattr(self.request.user, 'role') and self.request.user.role:
            if self.request.user.role.name == 'super_admin':
                # Super admin can see all teachers
                return queryset
        
        # Get school_id for filtering
        school_id = self.get_school_id()
        
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        logger.debug(f"TeacherViewSet - User: {self.request.user.username}, School ID: {school_id}")
        
        # For non-super-admin users, filter by school_id
        if school_id:
            # Filter by school_id
            queryset = queryset.filter(school_id=school_id)
            logger.debug(f"Filtered teachers by school_id={school_id}, count: {queryset.count()}")
        else:
            logger.warning(f"No school_id found for user {self.request.user.username}")
            # If no school_id found for authenticated user, return empty queryset
            return queryset.none()
        
        return queryset


class StudentViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Student management"""
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['school', 'applying_class', 'category', 'gender', 'student_id']
    search_fields = ['student_name', 'parent_name', 'admission_number', 'email']
    ordering_fields = ['created_at', 'student_name']
    ordering = ['-created_at']

    def get_permissions(self):
        """Require authentication for list/retrieve to ensure school filtering works"""
        # Changed: Require authentication for list/retrieve to enable school filtering
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]  # Changed from AllowAny() to ensure school filtering
        if self.action in ['create', 'destroy']:
            return [AllowAny()]  # Keep AllowAny for create/destroy if needed
        return [IsAuthenticated(), IsManagementAdmin()]
    
    def get_queryset(self):
        """
        Override to ensure school_id filtering is applied when user is authenticated.
        Now all list/retrieve requests are authenticated, so filtering will always work.
        """
        # Get base queryset
        queryset = super(SchoolFilterMixin, self).get_queryset()
        
        # All requests should be authenticated now (due to permission change above)
        # But keep the check for safety
        if not self.request.user.is_authenticated:
            return queryset.none()  # Changed: Return empty instead of all students
        
        # For authenticated users, apply school filtering via SchoolFilterMixin
        # Check if user is super admin (should see all data)
        if hasattr(self.request.user, 'role') and self.request.user.role:
            if self.request.user.role.name == 'super_admin':
                return queryset
        
        # Get school_id for current user
        school_id = self.get_school_id()
        
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        logger.debug(f"StudentViewSet - User: {self.request.user.username}, School ID: {school_id}")
        
        if not school_id:
            logger.warning(f"No school_id found for user {self.request.user.username}")
            # If no school_id found for authenticated user, return empty queryset
            return queryset.none()
        
        # Filter by school_id (Student model has 'school' ForeignKey)
        queryset = queryset.filter(school__school_id=school_id)
        logger.debug(f"Filtered students by school__school_id={school_id}, count: {queryset.count()}")
        
        return queryset
    
    def perform_create(self, serializer):
        """Override to ensure school is set correctly when creating students"""
        school_id = self.get_school_id()
        
        # Check if user is super admin
        if hasattr(self.request.user, 'role') and self.request.user.role:
            if self.request.user.role.name == 'super_admin':
                # Super admin can set school manually or leave it
                # If school is provided in data, use it; otherwise let it be set manually
                if 'school' not in serializer.validated_data:
                    # Try to get school from request data
                    school_id_from_request = self.request.data.get('school')
                    if school_id_from_request:
                        from super_admin.models import School
                        try:
                            school = School.objects.get(school_id=school_id_from_request)
                            serializer.save(school=school)
                            return
                        except School.DoesNotExist:
                            pass
                super().perform_create(serializer)
                return
        
        # For non-super-admin users, automatically set school
        if school_id:
            from super_admin.models import School
            try:
                school = School.objects.get(school_id=school_id)
                # Always set school, even if provided in data (to prevent cross-school data)
                serializer.save(school=school)
                return
            except School.DoesNotExist:
                pass
        
        # If no school_id found and user is authenticated, this is an error
        if self.request.user.is_authenticated:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({
                'school': 'No school associated with your account. Please contact administrator.'
            })
        
        # If user is not authenticated, try to use school from serializer if provided
        # Otherwise, let parent handle it (might fail validation)
        student = serializer.save()
        
        # Auto-link user if email matches and user exists
        if student.email:
            from main_login.models import User
            try:
                user = User.objects.get(email=student.email)
                if not student.user:
                    student.user = user
                    student.save()
            except User.DoesNotExist:
                pass
        
        return student


class NewAdmissionViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for New Admission management"""
    queryset = NewAdmission.objects.all()
    serializer_class = NewAdmissionSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'applying_class', 'category', 'gender', 'student_id']
    search_fields = ['student_name', 'parent_name', 'parent_phone', 'email', 'admission_number', 'student_id']
    ordering_fields = ['created_at', 'status', 'student_name']
    ordering = ['-created_at']
    
    def create(self, request, *args, **kwargs):
        """Override create to provide better error messages"""
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        
        # Log incoming request data for debugging
        logger.info(f"Admission creation request from user: {request.user.username}")
        logger.debug(f"Request data: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        
        if not serializer.is_valid():
            # Log validation errors
            logger.error(f"Validation failed for admission creation")
            logger.error(f"Validation errors: {serializer.errors}")
            logger.error(f"Request data received: {request.data}")
            
            return Response(
                {
                    'success': False,
                    'message': 'Validation error',
                    'errors': serializer.errors,
                    'received_data': {k: v for k, v in request.data.items()},  # Include received data for debugging
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            
            # Get generated password from context
            generated_password = serializer.context.get('generated_password', None)
            
            # Prepare response data
            response_data = serializer.data.copy()
            if generated_password:
                response_data['generated_password'] = generated_password
                response_data['login_credentials'] = {
                    'email': serializer.data.get('email'),
                    'password': generated_password,
                    'message': 'Please save these credentials. You can use them to login once admission is approved.'
                }
            
            logger.info(f"Admission created successfully: {serializer.data.get('student_id')}")
            return Response(
                {
                    'success': True,
                    'message': 'Admission created successfully. User account created with generated password.',
                    'data': response_data,
                },
                status=status.HTTP_201_CREATED,
                headers=headers
            )
        except Exception as e:
            # Log exception details
            logger.error(f"Exception during admission creation: {str(e)}")
            logger.error(traceback.format_exc())
            
            return Response(
                {
                    'success': False,
                    'message': str(e),
                    'error': 'Failed to create admission',
                    'error_type': type(e).__name__,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def perform_create(self, serializer):
        """Override to handle admission creation - user creation is done in serializer"""
        # The serializer's create method already handles user creation
        # Just save the admission (which will call serializer.create())
        admission = serializer.save()
        
        # Get generated password from serializer if set (the serializer's create method sets this)
        generated_password = getattr(serializer, 'generated_password', None)
        if generated_password:
            serializer.context['generated_password'] = generated_password
        
        return admission
    
    def update(self, request, *args, **kwargs):
        """Override update to provide better error messages and handle partial updates"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        old_status = instance.status
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        
        if not serializer.is_valid():
            return Response(
                {
                    'success': False,
                    'message': 'Validation error',
                    'errors': serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get the new status from validated data
            new_status = serializer.validated_data.get('status', old_status)
            
            # Perform the update
            self.perform_update(serializer)
            
            # Refresh instance to get updated data
            instance.refresh_from_db()
            
            # If status changed to 'Approved', create Student record
            created_student = None
            if old_status != 'Approved' and new_status == 'Approved':
                try:
                    created_student = instance.create_student_from_admission()
                except Exception as e:
                    return Response(
                        {
                            'success': False,
                            'message': f'Admission approved but failed to create student record: {str(e)}',
                            'error': 'Failed to create student',
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            if getattr(instance, '_prefetched_objects_cache', None):
                instance._prefetched_objects_cache = {}
            
            # Prepare response data
            response_data = serializer.data.copy()
            if created_student:
                # Include student information in response
                student_serializer = StudentSerializer(created_student)
                response_data['created_student'] = student_serializer.data
                message = 'Admission approved and student record created successfully'
            else:
                message = 'Admission updated successfully'
            
            return Response(
                {
                    'success': True,
                    'message': message,
                    'data': response_data,
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    'success': False,
                    'message': str(e),
                    'error': 'Failed to update admission',
                },
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """
        Custom action to approve an admission and create Student record.
        POST /api/management-admin/admissions/{id}/approve/
        """
        admission = self.get_object()
        
        if admission.status == 'Approved':
            return Response(
                {
                    'success': False,
                    'message': 'Admission is already approved',
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update status to Approved
        old_status = admission.status
        admission.status = 'Approved'
        
        # Generate admission number if not provided
        if not admission.admission_number:
            import datetime
            counter = 0
            max_attempts = 1000  # Prevent infinite loops
            while counter < max_attempts:
                # Use current timestamp with microseconds and add random component for uniqueness
                now = datetime.datetime.now()
                timestamp = now.strftime('%Y%m%d%H%M%S%f')
                random_suffix = ''.join(random.choices(string.digits, k=3))  # Add 3 random digits
                admission_number = f'ADM-{now.year}-{timestamp[-9:]}{random_suffix}'  # Use last 9 chars + 3 random
                
                # Check uniqueness
                if not Student.objects.filter(admission_number=admission_number).exists() and \
                   not NewAdmission.objects.filter(admission_number=admission_number).exists():
                    break
                counter += 1
            
            if counter >= max_attempts:
                # Fallback: use timestamp with milliseconds and random string
                now = datetime.datetime.now()
                random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
                admission_number = f'ADM-{now.year}-{now.strftime("%m%d%H%M%S")}{random_str}'
            
            admission.admission_number = admission_number
        
        admission.save()
        
        # Create Student record
        created_student = None
        try:
            created_student = admission.create_student_from_admission()
        except Exception as e:
            # Rollback status if student creation fails
            admission.status = old_status
            admission.save()
            return Response(
                {
                    'success': False,
                    'message': f'Failed to create student record: {str(e)}',
                    'error': 'Student creation failed',
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Serialize response
        serializer = self.get_serializer(admission)
        response_data = serializer.data.copy()
        
        if created_student:
            student_serializer = StudentSerializer(created_student)
            response_data['created_student'] = student_serializer.data
        
        return Response(
            {
                'success': True,
                'message': 'Admission approved and student record created successfully',
                'data': response_data,
            },
            status=status.HTTP_200_OK
        )
    


class DashboardViewSet(viewsets.ViewSet):
    """ViewSet for Dashboard data"""
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get dashboard statistics"""
        # Get school associated with the user (assuming user has a school)
        # This would need to be implemented based on your user-school relationship
        stats = DashboardStats.objects.first()  # Placeholder
        
        if stats:
            serializer = DashboardStatsSerializer(stats)
            return Response(serializer.data)
        
        return Response({
            'total_teachers': 0,
            'total_students': 0,
            'total_departments': 0,
        })


class ExaminationManagementViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Examination Management"""
    queryset = Examination_management.objects.all()
    serializer_class = ExaminationManagementSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['Exam_Type', 'Exam_Status']
    search_fields = ['Exam_Title', 'Exam_Description', 'Exam_Location']
    ordering_fields = ['Exam_Date', 'Exam_Created_At', 'Exam_Title']
    ordering = ['-Exam_Created_At']
    
    def get_permissions(self):
        """Allow read/create/update/delete without auth for development - can be adjusted"""
        if self.action in ['list', 'retrieve', 'create', 'update', 'partial_update', 'destroy']:
            return [AllowAny()]
        return [IsAuthenticated(), IsManagementAdmin()]


class FeeViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Fee Management"""
    queryset = Fee.objects.select_related('student').prefetch_related('payment_history').all()
    serializer_class = FeeSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['fee_type', 'status', 'frequency', 'grade', 'student']
    search_fields = ['student__student_name', 'description', 'fee_type']
    ordering_fields = ['due_date', 'created_at', 'total_amount']
    ordering = ['-due_date']
    
    def get_permissions(self):
        """Allow read/create/update/delete without auth for development - can be adjusted"""
        if self.action in ['list', 'retrieve', 'create', 'update', 'partial_update', 'destroy']:
            return [AllowAny()]
        return [IsAuthenticated(), IsManagementAdmin()]
    
    def get_queryset(self):
        """Override to ensure school_id filtering is applied"""
        queryset = super().get_queryset()
        
        # Check if user is super admin
        if hasattr(self.request.user, 'role') and self.request.user.role:
            if self.request.user.role.name == 'super_admin':
                return queryset
        
        # Get school_id for filtering
        school_id = self.get_school_id()
        if school_id:
            # Filter by school_id
            queryset = queryset.filter(school_id=school_id)
        else:
            # If no school_id and user is authenticated (not super admin), return empty
            if self.request.user.is_authenticated:
                return queryset.none()
        
        return queryset
    
    def filter_queryset(self, queryset):
        """Override filter_queryset to handle student filter by student_id_string"""
        # Handle student filter by student_id_string if provided
        student_param = self.request.query_params.get('student')
        if student_param:
            # Try to filter by student_id_string first (e.g., STUD-003)
            queryset = queryset.filter(student_id_string=student_param)
            # If no results, try filtering by student__student_id
            if not queryset.exists():
                queryset = self.get_queryset().filter(student__student_id=student_param)
            # Remove 'student' from query params to prevent ForeignKey filter error
            # Create a mutable copy of query params
            from django.http import QueryDict
            mutable_params = self.request.query_params.copy()
            mutable_params._mutable = True
            mutable_params.pop('student', None)
            mutable_params._mutable = False
            # Temporarily replace query_params
            original_get = self.request._request.GET
            self.request._request.GET = mutable_params
            try:
                # Call parent filter_queryset without student param
                queryset = super().filter_queryset(queryset)
            finally:
                self.request._request.GET = original_get
        else:
            queryset = super().filter_queryset(queryset)
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """Override list to ensure proper response format and handle errors gracefully"""
        try:
            # Get queryset (already filtered by school_id via get_queryset)
            queryset = self.get_queryset()
            
            # Apply filters (student filter is handled in filter_queryset override)
            queryset = self.filter_queryset(queryset)
            
            # Serialize the data
            serializer = self.get_serializer(queryset, many=True)
            data = serializer.data
            
            return Response(data)
        except Exception as e:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in FeeViewSet.list: {e}")
            logger.error(traceback.format_exc())
            
            return Response(
                {'error': str(e), 'detail': 'An error occurred while fetching fees'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='record-payment')
    def record_payment(self, request, pk=None):
        """Record a payment for a fee and create payment history"""
        try:
            fee = self.get_object()
            payment_amount = request.data.get('payment_amount')
            payment_date = request.data.get('payment_date')
            receipt_number = request.data.get('receipt_number', '')
            notes = request.data.get('notes', '')
            
            if not payment_amount:
                return Response(
                    {'error': 'payment_amount is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            try:
                payment_amount = float(payment_amount)
            except (ValueError, TypeError):
                return Response(
                    {'error': 'payment_amount must be a valid number'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Parse payment date or use today
            from datetime import date
            if payment_date:
                try:
                    from datetime import datetime
                    payment_date_obj = datetime.strptime(payment_date, '%Y-%m-%d').date()
                except ValueError:
                    payment_date_obj = date.today()
            else:
                payment_date_obj = date.today()
            
            # Create payment history record
            from django.utils import timezone
            payment_history = PaymentHistory.objects.create(
                fee=fee,
                payment_amount=payment_amount,
                payment_date=payment_date_obj,
                receipt_number=receipt_number,
                notes=notes
            )
            print(f"Created payment history: ID={payment_history.id}, Amount={payment_history.payment_amount}, Date={payment_history.payment_date}, Created={payment_history.created_at}")
            
            # If receipt was uploaded first (before payment), link it now
            # Check if fee has an uploaded receipt with matching receipt number
            if receipt_number and fee.upload_receipt:
                # Link the receipt to this payment history entry
                payment_history.upload_receipt = fee.upload_receipt
                payment_history.save()
                # Clear fee.upload_receipt as it's now linked to payment history
                fee.upload_receipt = ''
                fee.save()
            
            # Update fee with new payment (cumulative)
            # Set last_paid_date (will be set even for first payment)
            fee.last_paid_date = payment_date_obj
            
            # Update paid_amount cumulatively
            from decimal import Decimal
            fee.paid_amount = Decimal(str(fee.paid_amount)) + Decimal(str(payment_amount))
            
            # Recalculate due amount (will be recalculated in save() method too)
            fee.due_amount = Decimal(str(fee.total_amount)) - Decimal(str(fee.paid_amount))
            
            # Update status (using Decimal comparison)
            if fee.paid_amount >= fee.total_amount:
                fee.status = 'paid'
            elif fee.paid_amount > Decimal('0'):
                fee.status = 'pending'
            
            # Save the fee (this will trigger the save() method which recalculates due_amount)
            fee.save()
            print(f"Fee saved: ID={fee.id}, paid_amount={fee.paid_amount}, due_amount={fee.due_amount}, last_paid_date={fee.last_paid_date}, status={fee.status}")
            
            # Reload from database to ensure we have the latest data including payment history
            fee = Fee.objects.prefetch_related('payment_history').select_related('student').get(pk=fee.pk)
            
            # Verify the calculations
            print(f"Fee after refresh from database:")
            print(f"  - total_amount: {fee.total_amount}")
            print(f"  - paid_amount: {fee.paid_amount}")
            print(f"  - due_amount: {fee.due_amount}")
            print(f"  - last_paid_date: {fee.last_paid_date}")
            print(f"  - calculated due: {float(fee.total_amount) - float(fee.paid_amount)}")
            print(f"  - payment_history count: {fee.payment_history.count()}")
            
            # Return updated fee with payment history
            serializer = self.get_serializer(fee)
            serialized_data = serializer.data
            print(f"Serialized data:")
            print(f"  - paid_amount: {serialized_data.get('paid_amount')}")
            print(f"  - due_amount: {serialized_data.get('due_amount')}")
            print(f"  - last_paid_date: {serialized_data.get('last_paid_date')}")
            print(f"  - payment_history items: {len(serialized_data.get('payment_history', []))}")
            return Response(serialized_data, status=status.HTTP_200_OK)
            
        except Fee.DoesNotExist:
            return Response(
                {'error': 'Fee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            print(f"Error recording payment: {e}")
            print(traceback.format_exc())
            return Response(
                {'error': str(e), 'detail': 'An error occurred while recording payment'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='upload-receipt')
    def upload_receipt(self, request, pk=None):
        """Upload receipt for a fee with receipt number (same as teacher profile photo)"""
        try:
            fee = self.get_object()
            receipt_number = request.data.get('receipt_number', '')
            
            if not receipt_number:
                return Response(
                    {'error': 'receipt_number is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Handle receipt file upload (same as teacher profile photo)
            receipt_file = request.FILES.get('receipt_file')
            receipt_url = None
            
            if receipt_file:
                # Save file and get the URL/path (same as teacher profile photo)
                from django.core.files.storage import default_storage
                from django.utils import timezone
                import os
                
                # Generate filename with timestamp to avoid conflicts
                file_ext = os.path.splitext(receipt_file.name)[1]
                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                filename = f'receipts/{timestamp}_{receipt_number}_{receipt_file.name}'
                
                # Save file using default storage
                saved_path = default_storage.save(filename, receipt_file)
                # Get the URL for the saved file
                receipt_url = default_storage.url(saved_path)
            
            if not receipt_url:
                return Response(
                    {'error': 'receipt_file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if payment history with this receipt number already exists
            # (user marked as paid first, then uploading receipt)
            if receipt_number:
                payment_history = PaymentHistory.objects.filter(
                    fee=fee,
                    receipt_number=receipt_number
                ).first()
                
                if payment_history:
                    # Payment history exists - link receipt directly
                    payment_history.upload_receipt = receipt_url
                    payment_history.save()
                else:
                    # Payment history doesn't exist yet - store in fee temporarily
                    # It will be linked when payment is recorded with matching receipt_number
                    fee.upload_receipt = receipt_url
                    fee.save()
            
            # Return updated fee
            serializer = self.get_serializer(fee)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Fee.DoesNotExist:
            return Response(
                {'error': 'Fee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            print(f"Error uploading receipt: {e}")
            print(traceback.format_exc())
            return Response(
                {'error': str(e), 'detail': 'An error occurred while uploading receipt'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['put', 'patch'], url_path='payment-history/(?P<payment_id>[^/.]+)')
    def update_payment_history(self, request, pk=None, payment_id=None):
        """Update a payment history record and recalculate fee totals"""
        try:
            fee = self.get_object()
            
            # Get the payment history record
            try:
                payment_history = PaymentHistory.objects.get(id=payment_id, fee=fee)
            except PaymentHistory.DoesNotExist:
                return Response(
                    {'error': 'Payment history record not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Get old payment amount before update
            old_amount = float(payment_history.payment_amount)
            
            # Update payment history fields
            if 'payment_amount' in request.data:
                try:
                    new_amount = float(request.data.get('payment_amount'))
                    payment_history.payment_amount = new_amount
                except (ValueError, TypeError):
                    return Response(
                        {'error': 'payment_amount must be a valid number'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            if 'payment_date' in request.data:
                from datetime import date, datetime
                payment_date = request.data.get('payment_date')
                if payment_date:
                    try:
                        payment_date_obj = datetime.strptime(payment_date, '%Y-%m-%d').date()
                        payment_history.payment_date = payment_date_obj
                    except ValueError:
                        pass  # Keep existing date if invalid
            
            if 'receipt_number' in request.data:
                payment_history.receipt_number = request.data.get('receipt_number', '')
            
            if 'notes' in request.data:
                payment_history.notes = request.data.get('notes', '')
            
            payment_history.save()
            print(f"Updated payment history: ID={payment_history.id}, Amount={payment_history.payment_amount}, Date={payment_history.payment_date}")
            
            # Refresh fee from database to get updated payment_history
            fee.refresh_from_db()
            fee = Fee.objects.prefetch_related('payment_history').select_related('student').get(pk=fee.pk)
            
            # Recalculate fee's paid_amount by summing all payment history
            from decimal import Decimal
            total_paid = Decimal('0')
            payment_count = 0
            for payment in fee.payment_history.all():
                total_paid += Decimal(str(payment.payment_amount))
                payment_count += 1
                print(f"Payment {payment_count}: ID={payment.id}, Amount={payment.payment_amount}")
            
            print(f"Total paid calculated: {total_paid}, from {payment_count} payments")
            fee.paid_amount = total_paid
            
            # Recalculate due amount
            fee.due_amount = Decimal(str(fee.total_amount)) - fee.paid_amount
            print(f"Fee amounts - Total: {fee.total_amount}, Paid: {fee.paid_amount}, Due: {fee.due_amount}")
            
            # Update last_paid_date to the most recent payment date
            latest_payment = fee.payment_history.order_by('-payment_date').first()
            if latest_payment:
                fee.last_paid_date = latest_payment.payment_date
                print(f"Last paid date updated to: {fee.last_paid_date}")
            
            # Update status
            if fee.paid_amount >= fee.total_amount:
                fee.status = 'paid'
            elif fee.paid_amount > Decimal('0'):
                fee.status = 'pending'
            else:
                fee.status = 'pending'
            
            fee.save()
            print(f"Fee saved: ID={fee.id}, paid_amount={fee.paid_amount}, due_amount={fee.due_amount}, status={fee.status}")
            
            # Reload from database one more time to ensure we have the latest data
            fee = Fee.objects.prefetch_related('payment_history').select_related('student').get(pk=fee.pk)
            
            # Return updated fee with payment history
            serializer = self.get_serializer(fee)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Fee.DoesNotExist:
            return Response(
                {'error': 'Fee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            print(f"Error updating payment history: {e}")
            print(traceback.format_exc())
            return Response(
                {'error': str(e), 'detail': 'An error occurred while updating payment history'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='student-summary')
    def student_summary(self, request):
        """Get fee summary grouped by student"""
        try:
            student_id = request.query_params.get('student_id')
            if not student_id:
                return Response(
                    {'error': 'student_id parameter is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get all fees for this student
            queryset = self.get_queryset()
            fees = queryset.filter(
                student_id_string=student_id
            ) | queryset.filter(
                student__student_id=student_id
            )
            
            if not fees.exists():
                return Response(
                    {'error': 'No fees found for this student'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Get student info from first fee
            first_fee = fees.first()
            student = first_fee.student
            
            # Calculate totals
            total_payable = sum(float(fee.total_amount) for fee in fees)
            total_paid = sum(float(fee.paid_amount) for fee in fees)
            total_due = sum(float(fee.due_amount) for fee in fees)
            
            # Get next due date (earliest unpaid fee)
            unpaid_fees = fees.filter(status__in=['pending', 'overdue']).order_by('due_date')
            next_due_date = unpaid_fees.first().due_date if unpaid_fees.exists() else None
            
            # Determine overall payment status
            if total_due == 0:
                payment_status = 'FULLY PAID'
            elif total_paid > 0:
                payment_status = 'PARTIALLY PAID'
            else:
                payment_status = 'NOT PAID'
            
            # Group fees by fee_type
            fees_by_type = {}
            for fee in fees:
                fee_type = fee.fee_type
                if fee_type not in fees_by_type:
                    fees_by_type[fee_type] = []
                fees_by_type[fee_type].append(self.get_serializer(fee).data)
            
            # Get all payment history (sorted by date, newest first)
            all_payment_history = []
            for fee in fees:
                for payment in fee.payment_history.all():
                    all_payment_history.append({
                        'id': payment.id,
                        'fee_type': fee.fee_type,
                        'payment_amount': float(payment.payment_amount),
                        'payment_date': payment.payment_date.isoformat(),
                        'receipt_number': payment.receipt_number,
                        'upload_receipt': payment.upload_receipt if payment.upload_receipt else None,
                        'notes': payment.notes,
                        'created_at': payment.created_at.isoformat() if payment.created_at else None,
                    })
            
            # Sort payment history by date (newest first)
            all_payment_history.sort(key=lambda x: x['payment_date'], reverse=True)
            
            # Get student grade if available
            student_grade = ''
            if hasattr(student, 'grade'):
                student_grade = student.grade
            elif hasattr(student, 'student_grade'):
                student_grade = student.student_grade
            
            return Response({
                'student': {
                    'student_id': str(student.student_id) if hasattr(student, 'student_id') else student_id,
                    'student_name': student.student_name,
                    'applying_class': student.applying_class,
                    'grade': student_grade,
                    'email': student.email,
                },
                'summary': {
                    'total_payable': total_payable,
                    'total_paid': total_paid,
                    'total_due': total_due,
                    'next_due_date': next_due_date.isoformat() if next_due_date else None,
                    'payment_status': payment_status,
                },
                'fees_by_type': fees_by_type,
                'payment_history': all_payment_history,
            })
        except Exception as e:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in student_summary: {e}")
            logger.error(traceback.format_exc())
            return Response(
                {'error': str(e), 'detail': 'An error occurred while fetching student fee summary'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class BusViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Bus management"""
    queryset = Bus.objects.all()
    serializer_class = BusSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    lookup_field = 'bus_number'  # Use bus_number as primary key for lookups
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['school', 'bus_type', 'is_active']
    search_fields = ['bus_number', 'driver_name', 'route_name', 'registration_number']
    ordering_fields = ['bus_number', 'created_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Optimize queryset with prefetch_related for stops and students"""
        queryset = super().get_queryset()
        # Prefetch stops and their students to avoid N+1 queries
        queryset = queryset.prefetch_related(
            'stops__stop_students__student',
            'stops__bus'
        ).select_related('school')
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Override create to automatically get school from logged-in user if not provided"""
        # Get school from request data if provided
        school_id = request.data.get('school')
        
        # If school_id not provided, try to get it from the logged-in user
        if not school_id:
            try:
                school = School.objects.filter(user=request.user).first()
                if not school:
                    return Response(
                        {
                            'success': False,
                            'message': 'No school found. Please contact administrator to assign a school to your account.',
                            'error': 'School not found'
                        },
                        status=status.HTTP_404_NOT_FOUND
                    )
                # Create mutable copy of request data and add school_id
                mutable_data = request.data.copy()
                if hasattr(mutable_data, '_mutable'):
                    mutable_data._mutable = True
                mutable_data['school'] = school.school_id
                # Update request data
                request._full_data = mutable_data
            except Exception as e:
                return Response(
                    {
                        'success': False,
                        'message': f'Error getting school: {str(e)}',
                        'error': 'Failed to get school'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Call parent create method
        return super().create(request, *args, **kwargs)


class BusStopViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for BusStop management"""
    queryset = BusStop.objects.all()
    serializer_class = BusStopSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    lookup_field = 'stop_id'  # Explicitly set lookup field to stop_id (CharField: busnumber_stopnumber)
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['route_type']  # Removed 'bus' from here, will handle manually
    search_fields = ['stop_name']
    ordering_fields = ['stop_order', 'stop_time']
    ordering = ['bus', 'route_type', 'stop_order']
    
    def get_queryset(self):
        """Override to handle bus filter by bus_number instead of bus_id"""
        queryset = super().get_queryset()
        
        # Handle bus filter - if bus parameter is provided, filter by bus__bus_number
        bus_param = self.request.query_params.get('bus')
        if bus_param:
            queryset = queryset.filter(bus__bus_number=bus_param)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def students(self, request, *args, **kwargs):
        """Get all students for a specific stop.
        For afternoon stops, returns students from the corresponding morning stop (matched by stop_name).
        """
        # get_object() automatically uses the lookup_field (stop_id) to find the object
        # Using *args, **kwargs to avoid parameter name conflicts with custom lookup_field
        stop = self.get_object()
        
        # If this is an afternoon stop, get students from corresponding morning stop
        if stop.route_type == 'afternoon':
            # Find corresponding morning stop with same stop_name
            try:
                corresponding_morning_stop = BusStop.objects.filter(
                    bus=stop.bus,
                    route_type='morning',
                    stop_name=stop.stop_name
                ).first()
                
                if corresponding_morning_stop:
                    # Return students from corresponding morning stop
                    students = corresponding_morning_stop.stop_students.all()
                else:
                    # Fallback to afternoon stop's own students if no corresponding morning stop
                    students = stop.stop_students.all()
            except Exception:
                # Fallback to afternoon stop's own students on error
                students = stop.stop_students.all()
        else:
            # For morning stops, return students directly assigned to this stop
            students = stop.stop_students.all()
        
        serializer = BusStopStudentSerializer(students, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], url_path='drop-off-view')
    def drop_off_view(self, request, pk=None):
        """
        Get drop-off view for the first morning stop.
        Returns the last drop-off stop information if this is the first morning stop.
        """
        stop = self.get_object()
        
        # Check if this is the first morning stop
        if stop.route_type == 'morning' and stop.stop_order == 1:
            drop_off_stop = stop.corresponding_drop_off_stop
            if drop_off_stop:
                serializer = BusStopSerializer(drop_off_stop)
                return Response({
                    'success': True,
                    'message': 'Drop-off stop information for first morning stop',
                    'first_morning_stop': BusStopSerializer(stop).data,
                    'last_drop_off_stop': serializer.data
                })
            else:
                return Response({
                    'success': False,
                    'message': 'No drop-off stop found. The last drop-off stop will be created automatically when saved.',
                    'first_morning_stop': BusStopSerializer(stop).data
                })
        else:
            return Response({
                'success': False,
                'message': 'This endpoint is only available for the first morning stop (stop_order=1, route_type=morning)'
            }, status=status.HTTP_400_BAD_REQUEST)


class BusStopStudentViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for BusStopStudent management"""
    queryset = BusStopStudent.objects.all()
    serializer_class = BusStopStudentSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['bus_stop', 'student']
    search_fields = ['student_name', 'student_id_string', 'student__student_id', 'student__student_name']
    ordering_fields = ['created_at']
    ordering = ['bus_stop', 'student_name']
    
    def create(self, request, *args, **kwargs):
        """Override create to fetch student by student_id and validate school match"""
        student_id = request.data.get('student_id')
        stop_id = request.data.get('stop')
        
        if not student_id:
            return Response(
                {'error': 'student_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not stop_id:
            return Response(
                {'error': 'stop is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            stop = BusStop.objects.select_related('bus', 'bus__school').get(pk=stop_id)
        except BusStop.DoesNotExist:
            return Response(
                {'error': 'Stop not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get the bus and its school
        bus = stop.bus
        bus_school = bus.school
        
        if not bus_school:
            return Response(
                {'error': 'Bus does not have an associated school'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Find student by student_id and filter by school_id to ensure it belongs to the same school
            student = Student.objects.select_related('school').get(
                student_id=student_id,
                school=bus_school
            )
        except Student.DoesNotExist:
            return Response(
                {
                    'error': f'Student with ID {student_id} not found in school {bus_school.school_name}. Please ensure the student belongs to the same school as the bus.',
                    'school_id': str(bus_school.school_id)
                },
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Additional validation: ensure student's school matches bus's school
        if student.school != bus_school:
            return Response(
                {
                    'error': f'Student belongs to a different school. Bus belongs to {bus_school.school_name}, but student belongs to {student.school.school_name}',
                    'bus_school_id': str(bus_school.school_id),
                    'student_school_id': str(student.school.school_id)
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if student is already assigned to this stop
        if BusStopStudent.objects.filter(bus_stop=stop, student=student).exists():
            return Response(
                {'error': 'Student is already assigned to this stop'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if student is already assigned to any other stop of the same bus
        existing_assignment = BusStopStudent.objects.filter(
            bus_stop__bus=bus,
            student=student
        ).select_related('bus_stop').first()
        
        if existing_assignment:
            return Response(
                {
                    'error': f'Student is already assigned to another stop: {existing_assignment.bus_stop.stop_name} (Stop {existing_assignment.bus_stop.stop_order}, {existing_assignment.bus_stop.get_route_type_display()})',
                    'existing_stop_name': existing_assignment.bus_stop.stop_name,
                    'existing_stop_order': existing_assignment.bus_stop.stop_order,
                    'existing_route_type': existing_assignment.bus_stop.route_type,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if student is already assigned to a DIFFERENT bus
        different_bus_assignment = BusStopStudent.objects.filter(
            student=student
        ).exclude(
            bus_stop__bus=bus
        ).select_related('bus_stop', 'bus_stop__bus').first()
        
        if different_bus_assignment:
            assigned_bus = different_bus_assignment.bus_stop.bus
            return Response(
                {
                    'error': f'Student is already assigned to Bus Number: {assigned_bus.bus_number}. Please remove the student from that bus first before assigning to this bus.',
                    'assigned_bus_number': assigned_bus.bus_number,
                    'assigned_stop_name': different_bus_assignment.bus_stop.stop_name,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create the BusStopStudent
        bus_stop_student = BusStopStudent.objects.create(
            bus_stop=stop,
            student=student
        )
        
        serializer = self.get_serializer(bus_stop_student)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class SchoolViewSet(viewsets.ViewSet):
    """ViewSet for getting current user's school"""
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    
    @action(detail=False, methods=['get'], url_path='current')
    def current(self, request):
        """Get the school associated with the current logged-in user"""
        from super_admin.models import School
        from super_admin.serializers import SchoolSerializer
        
        try:
            # Get school from user's school_account relationship
            school = School.objects.filter(user=request.user).first()
            
            if not school:
                return Response(
                    {
                        'success': False,
                        'message': 'No school found for this user. Please contact administrator.',
                        'error': 'School not found'
                    },
                    status=status.HTTP_404_NOT_FOUND
                )
            
            serializer = SchoolSerializer(school)
            return Response(
                {
                    'success': True,
                    'data': serializer.data
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    'success': False,
                    'message': str(e),
                    'error': 'Failed to fetch school'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class EventViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Event management"""
    queryset = Event.objects.all()
    serializer_class = EventSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'status', 'date']
    search_fields = ['name', 'location', 'organizer', 'description']
    ordering_fields = ['date', 'created_at', 'name']
    ordering = ['-date', '-created_at']
    
    def get_permissions(self):
        """Allow read/create/update/delete without auth for development - can be adjusted"""
        if self.action in ['list', 'retrieve', 'create', 'update', 'partial_update', 'destroy']:
            return [AllowAny()]
        return [IsAuthenticated(), IsManagementAdmin()]
    
    def create(self, request, *args, **kwargs):
        """Override create to automatically get school from logged-in user if not provided"""
        # Get school from request data if provided
        school_id = request.data.get('school')
        
        # If school_id not provided, try to get it from the logged-in user
        if not school_id:
            if request.user.is_authenticated:
                try:
                    school = School.objects.filter(user=request.user).first()
                    if not school:
                        # Try to get school_id from user's school relationship
                        school_id_from_user = get_user_school_id(request.user)
                        if school_id_from_user:
                            try:
                                school = School.objects.get(school_id=school_id_from_user)
                            except School.DoesNotExist:
                                school = None
                    
                    if school:
                        # Create mutable copy of request data and add school_id
                        mutable_data = request.data.copy()
                        if hasattr(mutable_data, '_mutable'):
                            mutable_data._mutable = True
                        mutable_data['school'] = school.school_id
                        # Update request data - use QueryDict for proper handling
                        from django.http import QueryDict
                        if isinstance(request.data, QueryDict):
                            request._full_data = QueryDict('', mutable=True)
                            request._full_data.update(mutable_data)
                        else:
                            request._full_data = mutable_data
                    else:
                        return Response(
                            {
                                'success': False,
                                'message': 'No school found. Please contact administrator to assign a school to your account.',
                                'error': 'School not found'
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                except Exception as e:
                    return Response(
                        {
                            'success': False,
                            'message': f'Error getting school: {str(e)}',
                            'error': 'Failed to get school'
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                # For unauthenticated requests (development), try to get school from first available school
                # This is for development/testing only
                try:
                    school = School.objects.first()
                    if school:
                        mutable_data = request.data.copy()
                        if hasattr(mutable_data, '_mutable'):
                            mutable_data._mutable = True
                        mutable_data['school'] = school.school_id
                        # Update request data - use QueryDict for proper handling
                        from django.http import QueryDict
                        if isinstance(request.data, QueryDict):
                            request._full_data = QueryDict('', mutable=True)
                            request._full_data.update(mutable_data)
                        else:
                            request._full_data = mutable_data
                    else:
                        return Response(
                            {
                                'success': False,
                                'message': 'No school found in database. Please create a school first.',
                                'error': 'School required'
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                except Exception as e:
                    return Response(
                        {
                            'success': False,
                            'message': f'Error getting school: {str(e)}',
                            'error': 'Failed to get school'
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
        
        # Call parent create method
        try:
            # Use request._full_data if it was modified, otherwise use request.data
            data_to_use = getattr(request, '_full_data', None)
            if data_to_use is None:
                data_to_use = request.data
            
            # Convert QueryDict to dict if needed for better handling
            if hasattr(data_to_use, 'dict'):
                data_to_use = data_to_use.dict()
            
            serializer = self.get_serializer(data=data_to_use)
            if not serializer.is_valid():
                # Return validation errors
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Activity validation errors: {serializer.errors}')
                logger.error(f'Received data: {data_to_use}')
                return Response(
                    {
                        'success': False,
                        'message': 'Validation error',
                        'errors': serializer.errors,
                        'received_data': {k: v for k, v in data_to_use.items()},
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(
                {
                    'success': True,
                    'data': serializer.data
                },
                status=status.HTTP_201_CREATED,
                headers=headers
            )
        except Exception as e:
            # Return detailed error information
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error creating activity: {str(e)}')
            logger.error(traceback.format_exc())
            return Response(
                {
                    'success': False,
                    'message': f'Error creating activity: {str(e)}',
                    'error': str(e),
                    'traceback': traceback.format_exc() if settings.DEBUG else None
                },
                status=status.HTTP_400_BAD_REQUEST
            )
    
    
    def perform_create(self, serializer):
        """Set school_id when creating event"""
        event = serializer.save()
        
        # Set school_id after save (since it's read-only in serializer)
        school_id = self.get_school_id()
        if school_id:
            from super_admin.models import School
            Event.objects.filter(pk=event.pk).update(school_id=school_id)
            try:
                school = School.objects.get(school_id=school_id)
                Event.objects.filter(pk=event.pk).update(school_name=school.school_name)
            except School.DoesNotExist:
                pass
    
    def perform_update(self, serializer):
        """Update event - school_id should already be set"""
        serializer.save()


class DepartmentViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Department management"""
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['name', 'code']
    search_fields = ['name', 'code', 'head_name', 'email']
    ordering_fields = ['name', 'faculty_count', 'student_count']
    ordering = ['name']

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'create', 'update', 'partial_update', 'destroy']:
            return [AllowAny()]
        return [IsAuthenticated(), IsManagementAdmin()]

    def perform_create(self, serializer):
        """Set school reference when creating department"""
        department = serializer.save()
        
        school_id = self.get_school_id()
        if school_id:
            from super_admin.models import School
            try:
                school = School.objects.get(school_id=school_id)
                Department.objects.filter(pk=department.pk).update(
                    school=school,
                    school_name=school.school_name
                )
            except School.DoesNotExist:
                pass


class CampusFeatureViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for CampusFeature management"""
    queryset = CampusFeature.objects.all()
    serializer_class = CampusFeatureSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'status']
    search_fields = ['name', 'description', 'location']
    ordering_fields = ['name', 'category', 'date_added']
    ordering = ['-date_added']

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'create', 'update', 'partial_update', 'destroy']:
            return [AllowAny()]
        return [IsAuthenticated(), IsManagementAdmin()]

    def perform_create(self, serializer):
        """Set school reference when creating campus feature"""
        feature = serializer.save()
        
        school_id = self.get_school_id()
        if school_id:
            from super_admin.models import School
            try:
                school = School.objects.get(school_id=school_id)
                CampusFeature.objects.filter(pk=feature.pk).update(
                    school=school,
                    school_name=school.school_name
                )
            except School.DoesNotExist:
                pass


class AwardViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Award management"""
    queryset = Award.objects.all()
    serializer_class = AwardSerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'level', 'date']
    search_fields = ['title', 'recipient', 'description']
    ordering_fields = ['date', 'created_at', 'title']
    ordering = ['-date', '-created_at']
    
    def create(self, request, *args, **kwargs):
        """Override create to handle school assignment and validation"""
        # Get school from request data if provided
        school_id = request.data.get('school_id') or request.data.get('school')
        
        # If school_id not provided, try to get it from the logged-in user
        if not school_id:
            if request.user.is_authenticated:
                from main_login.utils import get_user_school_id
                school_id = get_user_school_id(request.user)
            
            # For development/unauthenticated, get first school
            if not school_id:
                from super_admin.models import School
                school = School.objects.first()
                if school:
                    school_id = school.school_id
        
        # Add school_id to data
        mutable_data = request.data.copy()
        if school_id:
            mutable_data['school_id'] = school_id
            
            # Also try to set school_name
            from super_admin.models import School
            try:
                school = School.objects.get(school_id=school_id)
                mutable_data['school_name'] = school.school_name
            except School.DoesNotExist:
                pass
        
        serializer = self.get_serializer(data=mutable_data)
        if serializer.is_valid():
            self.perform_create(serializer)
            return Response(
                {
                    'success': True,
                    'data': serializer.data
                },
                status=status.HTTP_201_CREATED
            )
        
        return Response(
            {
                'success': False,
                'message': 'Validation error',
                'errors': serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Generate Excel template for bulk award import"""
        wb = openpyxl.Workbook()
        
        # Template Sheet
        ws = wb.active
        ws.title = "Template"
        headers = [
            'Team Number', 'Award Title', 'Category', 'Student ID', 
            'Date (YYYY-MM-DD)', 'Description', 'Level', 'Presented By', 
            'Recipient Name', 'Award Document'
        ]
        ws.append(headers)
        
        # Removed sample data as requested
        
        # Instructions Sheet
        ws_inst = wb.create_sheet("Instructions")
        ws_inst.append(['Column', 'Description', 'Mandatory'])
        ws_inst.append(['Team Number', 'Group rows with the same number to create a single Award card for a team (use 1, 2, 3...)', 'No'])
        ws_inst.append(['Award Title', 'Official title of the award (First row of team is used)', 'Yes'])
        ws_inst.append(['Category', 'Academic, Sports, Arts, etc. (First row of team is used)', 'Yes'])
        ws_inst.append(['Student ID', 'Valid Student ID from school records', 'Yes'])
        ws_inst.append(['Date', 'Date of award in YYYY-MM-DD format (First row of team is used)', 'Yes'])
        ws_inst.append(['Description', 'Brief details about the achievement', 'No'])
        ws_inst.append(['Level', 'School, District, State, etc. (First row of team is used)', 'Yes'])
        ws_inst.append(['Presented By', 'Name of the person or organization', 'No'])
        ws_inst.append(['Recipient Name', 'Team Name or Student Name (Optional)', 'No'])
        ws_inst.append(['Award Document', 'Place/Paste the certificate image for THIS student in this row', 'No'])
        
        # Instructions for Images
        ws_inst.append([])
        ws_inst.append(['Note on Documents:', 'Place/Paste the award photo or certificate directly into the "Award Document" column. Each image will be automatically linked to the student in that row.'])
        
        # Set column widths
        for ws_iterator in [ws, ws_inst]:
            for column_cells in ws_iterator.columns:
                try:
                    length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                    ws_iterator.column_dimensions[column_cells[0].column_letter].width = length + 2
                except:
                    pass

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=awards_import_template.xlsx'
        return response

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import awards from Excel with optional images"""
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return Response({'success': False, 'message': 'No Excel file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        # Get school_id
        school_id = get_user_school_id(request.user)
        if not school_id:
             # Development fallback
             school = School.objects.first()
             if school: school_id = school.school_id
        
        if not school_id:
            return Response({'success': False, 'message': 'Could not determine school ID'}, status=status.HTTP_400_BAD_REQUEST)
            
        school_name = ""
        try:
            from super_admin.models import School
            school_obj = School.objects.get(school_id=school_id)
            school_name = school_obj.school_name
        except:
            pass

        from datetime import datetime
        print(f"\n--- IMPORT_EXCEL: [{datetime.now().strftime('%H:%M:%S')}] Starting Import Process ---", flush=True)
        print(f"User: {request.user}, School ID: {school_id}", flush=True)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            if "Template" not in wb.sheetnames:
                return Response({'success': False, 'message': 'Invalid Excel format. Could not find "Template" sheet.'}, status=status.HTTP_400_BAD_REQUEST)
            
            ws = wb["Template"]
            print(f"DEBUG: [v3] Import Request - File: {excel_file.name}, Sheet: Template. Images in ws._images: {len(ws._images)}", flush=True)
            
            # Map images to rows
            image_map = {}
            
            # Helper to extract bytes from openpyxl image
            def get_img_bytes(img):
                try:
                    # Method 1: Get from p_img (PIL Image if available)
                    if hasattr(img, 'p_img'):
                        from io import BytesIO
                        out = BytesIO()
                        img.p_img.save(out, format=img.p_img.format or 'PNG')
                        return out.getvalue()
                    
                    # Method 2: Get from _data property
                    if hasattr(img, '_data'):
                        return img._data()

                    # Method 3: Get from image or ref
                    img_src = getattr(img, 'ref', None) or getattr(img, 'image', None)
                    if img_src and hasattr(img_src, 'tobytes'):
                        return img_src.tobytes()
                    elif img_src and hasattr(img_src, 'read'):
                        img_src.seek(0)
                        return img_src.read()
                except Exception as ex:
                    print(f"DEBUG: Error extracting bytes: {str(ex)}", flush=True)
                return None

            # 1. Try standard openpyxl images
            if ws._images:
                for img_idx, img in enumerate(ws._images):
                    row = None
                    try:
                        if hasattr(img.anchor, '_from'):
                            row = img.anchor._from.row
                        elif hasattr(img.anchor, 'row'):
                            row = img.anchor.row
                    except:
                        pass
                    
                    if row is not None:
                        img_bytes = get_img_bytes(img)
                        if img_bytes:
                            image_map[row] = img_bytes
                            print(f"DEBUG: Mapped Image {img_idx} to Row index {row} via ws._images", flush=True)

            # 2. Fallback: Manually extract from ZIP if no images found or as extra check
            if not image_map:
                print(f"DEBUG: [v3] Starting ZIP extraction flow...", flush=True)
                try:
                    import zipfile
                    import xml.etree.ElementTree as ET
                    import os
                    
                    excel_file.seek(0)
                    with zipfile.ZipFile(excel_file, 'r') as z:
                        all_files = z.namelist()
                        
                        # Namespaces
                        ns = {
                            'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                            'rd': 'http://schemas.microsoft.com/office/spreadsheetml/2017/richdata',
                            'rdr': 'http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel'
                        }

                        # -- PART A: Find drawing-based images (Standard Excel) --
                        drawing_files = [f for f in all_files if 'xl/drawings/drawing' in f and f.endswith('.xml')]
                        for drawing_path in drawing_files:
                            try:
                                drawing_xml = z.read(drawing_path)
                                draw_root = ET.fromstring(drawing_xml)
                                draw_rel_path = f"xl/drawings/_rels/{drawing_path.split('/')[-1]}.rels"
                                if draw_rel_path not in all_files: continue
                                
                                dr_rels_root = ET.fromstring(z.read(draw_rel_path))
                                draw_rels_map = {}
                                for rel in dr_rels_root:
                                    target = rel.get('Target')
                                    if target:
                                        if target.startswith('../media/'):
                                            full_media_path = 'xl/media/' + target.split('/')[-1]
                                        elif target.startswith('media/'):
                                            full_media_path = 'xl/drawings/' + target
                                        else:
                                            full_media_path = 'xl/' + target.replace('../', '')
                                        draw_rels_map[rel.get('Id')] = full_media_path

                                for anchor in draw_root:
                                    from_elem = anchor.find('.//{*}from')
                                    if from_elem is not None:
                                        row_elem = from_elem.find('{*}row')
                                        if row_elem is not None:
                                            row_idx = int(row_elem.text)
                                            blip = anchor.find('.//{*}blip')
                                            if blip is not None:
                                                blip_rId = None
                                                for attr in blip.attrib:
                                                    if 'embed' in attr.lower():
                                                        blip_rId = blip.get(attr)
                                                        break
                                                if blip_rId and blip_rId in draw_rels_map:
                                                    media_path = draw_rels_map[blip_rId]
                                                    if media_path in all_files:
                                                        image_map[row_idx] = z.read(media_path)
                                                        print(f"DEBUG: Mapped standard image for row {row_idx}", flush=True)
                            except Exception as de:
                                print(f"DEBUG: Error in drawing fallback: {str(de)}", flush=True)

                        # -- PART B: Find Rich Data images (Image in Cell) --
                        if not image_map:
                            print(f"DEBUG: [v3] No standard images found. Attempting Rich Data (Image-in-cell) extraction...", flush=True)
                            try:
                                # 1. Find the "Template" sheet file
                                wb_xml = ET.fromstring(z.read('xl/workbook.xml'))
                                target_rid = None
                                for s in wb_xml.findall('.//main:sheet', ns):
                                    if s.attrib.get('name') == 'Template':
                                        target_rid = s.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                                        break
                                
                                if target_rid:
                                    wb_rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
                                    sheet_path = None
                                    for r in wb_rels.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                        if r.attrib.get('Id') == target_rid:
                                            sheet_path = 'xl/' + r.attrib.get('Target')
                                            break
                                    
                                    if sheet_path and sheet_path in all_files:
                                        # 2. Parse sheet to map Row -> vm index
                                        sheet_xml = ET.fromstring(z.read(sheet_path))
                                        vm_map = {} # row_idx -> vm_index
                                        for row_node in sheet_xml.findall('.//main:row', ns):
                                            r = int(row_node.attrib.get('r', '0')) - 1
                                            for c in row_node.findall('main:c', ns):
                                                vm = c.attrib.get('vm')
                                                if vm:
                                                    vm_map[r] = int(vm)
                                        
                                        if vm_map and 'xl/metadata.xml' in all_files:
                                            # 3. Resolve vm index -> rv index (metadata.xml)
                                            meta_xml = ET.fromstring(z.read('xl/metadata.xml'))
                                            rv_indices = {} # vm_idx -> rv_idx
                                            # valueMetadata/bk is index-based (vm="1" is first [0])
                                            for i, bk in enumerate(meta_xml.findall('.//main:valueMetadata/main:bk', ns)):
                                                rc = bk.find('main:rc', ns)
                                                if rc is not None and rc.attrib.get('t') == '1': # XLRICHVALUE
                                                    rv_indices[i + 1] = int(rc.attrib.get('v', '0'))
                                            
                                            if rv_indices and 'xl/richData/rdrichvalue.xml' in all_files:
                                                # 4. Resolve rv index -> rel index (rdrichvalue.xml)
                                                rdv_xml = ET.fromstring(z.read('xl/richData/rdrichvalue.xml'))
                                                rel_indices = {} # rv_idx -> rel_idx
                                                for i, rv in enumerate(rdv_xml.findall('rd:rv', ns)):
                                                    v_tags = rv.findall('rd:v', ns)
                                                    if v_tags:
                                                        # First v tag is usually the rel index
                                                        rel_indices[i] = int(v_tags[0].text)
                                                
                                                if rel_indices and 'xl/richData/richValueRel.xml' in all_files:
                                                    # 5. Resolve rel index -> rId (richValueRel.xml)
                                                    rdrel_xml = ET.fromstring(z.read('xl/richData/richValueRel.xml'))
                                                    rid_indices = {} # rel_idx -> rId
                                                    for i, rel in enumerate(rdrel_xml.findall('rdr:rel', ns)):
                                                        rid_indices[i] = rel.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                                                    
                                                    # 6. Resolve rId -> media path
                                                    rdrel_rels_path = 'xl/richData/_rels/richValueRel.xml.rels'
                                                    if rid_indices and rdrel_rels_path in all_files:
                                                        rdrels_root = ET.fromstring(z.read(rdrel_rels_path))
                                                        media_map = {}
                                                        for r in rdrels_root.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                                            target = r.attrib.get('Target')
                                                            if target:
                                                                # Target is relative to xl/richData/
                                                                m_path = os.path.normpath(os.path.join('xl/richData', target)).replace('\\', '/')
                                                                media_map[r.attrib.get('Id')] = m_path
                                                        
                                                        # Final Assembly
                                                        for row_idx, vm_idx in vm_map.items():
                                                            rv_idx = rv_indices.get(vm_idx)
                                                            if rv_idx is not None:
                                                                rel_idx = rel_indices.get(rv_idx)
                                                                if rel_idx is not None:
                                                                    rid = rid_indices.get(rel_idx)
                                                                    if rid:
                                                                        img_path = media_map.get(rid)
                                                                        if img_path and img_path in all_files:
                                                                            image_map[row_idx] = z.read(img_path)
                                                                            print(f"DEBUG: Mapped Rich Data Image for Row {row_idx}", flush=True)

                            except Exception as re:
                                print(f"DEBUG: Error in Rich Data extraction: {str(re)}", flush=True)

                except Exception as ze:
                    print(f"DEBUG: [v3] ZIP Fallback failed: {str(ze)}", flush=True)

            print(f"DEBUG: [v3] Image Mapping Complete. Entries: {len(image_map)}. Keys: {list(image_map.keys())}", flush=True)

            # iter_rows(min_row=2) means we start from the second row (index 1)

            from django.utils import timezone
            awards_to_create = []
            errors = []
            
            # iter_rows(min_row=2) means we start from the second row (index 1)
            # openpyxl row index for headers is 0. Data starts at 1.
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            # Group rows by Team Number
            # If Team Number is blank, treat each row as its own team (single award)
            groups = []
            import uuid
            
            for i, row in enumerate(rows):
                if not any(row): continue
                
                team_val = str(row[0]).strip() if row[0] is not None else ""
                
                # If no team number, use a unique ID to treat it as a single award
                actual_team_id = team_val if team_val else f"single_{uuid.uuid4().hex[:8]}"
                
                # Metadata for the award (from the first row of a group)
                # Team Number (0), Award Title (1), Category (2), Student ID (3), Date (4), 
                # Description (5), Level (6), Presented By (7), Recipient Name (8), Award Document (Legacy 9)
                
                row_data = {
                    'row_idx': i + 1, # abs_row_idx
                    'team_id': actual_team_id,
                    'title': row[1],
                    'category': row[2],
                    'student_id': str(row[3]).strip() if row[3] else "",
                    'date_val': row[4],
                    'desc': row[5],
                    'level': row[6],
                    'presented_by': row[7],
                    'recipient': row[8],
                }
                
                # Find if group already exists
                group = next((g for g in groups if g['team_id'] == actual_team_id and actual_team_id and not actual_team_id.startswith('single_')), None)
                
                if group:
                    group['rows'].append(row_data)
                else:
                    groups.append({
                        'team_id': actual_team_id,
                        'rows': [row_data]
                    })

            from django.utils import timezone
            from django.db import transaction
            
            success_count = 0
            errors = []

            with transaction.atomic():
                for group in groups:
                    first_row = group['rows'][0]
                    
                    # Validate all students in group
                    valid_students = []
                    group_errors = []
                    
                    for r_data in group['rows']:
                        sid = r_data['student_id']
                        if not sid:
                            group_errors.append(f"Row {r_data['row_idx'] + 1}: Student ID is mandatory")
                            continue
                            
                        # Split IDs if comma-separated in a single cell (legacy support)
                        id_strings = [s.strip() for s in sid.split(',') if s.strip()]
                        for s_id in id_strings:
                            student = Student.objects.filter(student_id=s_id, school_id=school_id).first()
                            if not student:
                                student = Student.objects.filter(student_id__iexact=s_id, school_id=school_id).first()
                                if not student:
                                    group_errors.append(f"Row {r_data['row_idx'] + 1}: Student ID {s_id} not found")
                                    continue
                            valid_students.append(student)
                    
                    if group_errors:
                        errors.extend(group_errors)
                        continue
                    
                    if not valid_students:
                        continue
                        
                    # Parse date from first row
                    date_val = first_row['date_val']
                    try:
                        if isinstance(date_val, str):
                            from django.utils.dateparse import parse_date
                            date_obj = parse_date(date_val)
                        elif hasattr(date_val, 'date'):
                            date_obj = date_val.date()
                        else:
                            date_obj = date_val
                        if not date_obj: date_obj = timezone.now().date()
                    except:
                        date_obj = timezone.now().date()

                    # Final recipient logic
                    final_recipient = first_row['recipient']
                    if not final_recipient:
                        # If multiple students and no team name, use comma list
                        final_recipient = ", ".join([s.student_name for s in valid_students[:3]])
                        if len(valid_students) > 3:
                            final_recipient += f" and {len(valid_students)-3} others"
                    
                    final_student_ids = ",".join(list(set([s.student_id for s in valid_students])))

                    # Create Award
                    save_award = Award(
                        school_id=school_id,
                        school_name=school_name,
                        title=first_row['title'] or "Award",
                        category=first_row['category'] or "Other",
                        recipient=final_recipient,
                        student_ids=final_student_ids,
                        date=date_obj,
                        description=first_row['desc'] or "",
                        level=first_row['level'] or "School",
                        presented_by=first_row['presented_by'] or "",
                        team_id=group['team_id'] if not group['team_id'].startswith('single_') else ""
                    )
                    save_award.save()
                    award = save_award
                    
                    # Create Certificates for each row in group
                    for r_data in group['rows']:
                        abs_row_idx = r_data['row_idx']
                        if abs_row_idx in image_map:
                            image_content = image_map[abs_row_idx]
                            import time
                            ts = int(time.time() * 1000)
                            # Create Certificate object
                            cert = AwardCertificate(
                                award=award,
                                student_id=r_data['student_id'],
                            )
                            # Save file
                            file_name = f"{r_data['student_id']}_{ts}_cert.png"
                            cert.document.save(file_name, ContentFile(image_content), save=True)
                            
                            # Also set as main document if it's the first row
                            if r_data == first_row:
                                award.document.save(file_name, ContentFile(image_content), save=False)
                                award.save()
                    
                    success_count += 1
            
            if errors:
                print(f"DEBUG: Import failed with {len(errors)} errors: {errors}", flush=True)
                return Response({'success': False, 'message': 'Import failed due to errors', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)
            
            if success_count > 0:
                print(f"DEBUG: Successfully imported {success_count} awards (grouped)", flush=True)
                return Response({'success': True, 'message': f'Successfully imported {success_count} awards'})
            
            return Response({'success': False, 'message': 'No valid award data found in Excel'})

        except Exception as e:
            import traceback
            traceback.print_exc() # Print to console for real-time debugging
            return Response({
                'success': False, 
                'message': f'Error processing Excel: {str(e)}', 
                'trace': traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def perform_create(self, serializer):
        """Set school reference when creating award"""
        serializer.save()

    def perform_update(self, serializer):
        """Update award"""
        serializer.save()


class ActivityViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Activity management"""
    queryset = Activity.objects.all()
    serializer_class = ActivitySerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'status', 'start_date', 'end_date']
    search_fields = ['name', 'instructor', 'location', 'description']
    ordering_fields = ['created_at', 'start_date', 'name']
    ordering = ['-created_at']
    
    def get_permissions(self):
        """Allow read/create/update/delete without auth for development - can be adjusted"""
        if self.action in ['list', 'retrieve', 'create', 'update', 'partial_update', 'destroy']:
            return [AllowAny()]
        return [IsAuthenticated(), IsManagementAdmin()]

    def perform_create(self, serializer):
        """Set school reference when creating activity"""
        activity = serializer.save()
        
        school_id = self.get_school_id()
        if school_id:
            from super_admin.models import School
            # Create a separate update query to set the foreign key directly by ID
            # This avoids needing to fetch the School object if we only have the ID
            Activity.objects.filter(pk=activity.pk).update(school_id=school_id)
            
            # Try to set school_name as well if possible
            try:
                school = School.objects.get(school_id=school_id)
                Activity.objects.filter(pk=activity.pk).update(school_name=school.school_name)
            except School.DoesNotExist:
                pass

    def perform_update(self, serializer):
        """Update activity"""
        serializer.save()


class GalleryViewSet(SchoolFilterMixin, viewsets.ModelViewSet):
    """ViewSet for Gallery management"""
    queryset = Gallery.objects.prefetch_related('images').all()
    serializer_class = GallerySerializer
    permission_classes = [IsAuthenticated, IsManagementAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'date', 'school_id']
    search_fields = ['title', 'description', 'photographer', 'location']
    ordering_fields = ['date', 'created_at']
    ordering = ['-date']
    
    def perform_create(self, serializer):
        """Override to handle image uploads if any"""
        gallery = serializer.save()
        
        # Handle multiple image uploads if present in request.FILES
        # Expecting keys like 'images[0]', 'images[1]' or just 'images' list
        if self.request.FILES:
            for key, file in self.request.FILES.items():
                if key.startswith('image'):
                    # Save each image
                    GalleryImage.objects.create(
                        gallery=gallery,
                        image=file  # Django handles file upload to MEDIA_ROOT
                    )

